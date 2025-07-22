"""
Template Editor - A tool for editing templates from Excel files
"""

import json
import sys  # Added missing import
from openpyxl import load_workbook
from typing import Dict, Tuple, Any
import os
import shutil
from datetime import datetime
from pathlib import Path

def get_base_path():
    """Get the base path for the application, works with PyInstaller"""
    if getattr(sys, 'frozen', False):
        # If the application is frozen (packaged by PyInstaller)
        if hasattr(sys, '_MEIPASS'):
            # PyInstaller one-file mode
            return os.path.join(sys._MEIPASS, 'Main')
        else:
            # PyInstaller one-dir mode
            return os.path.join(os.path.dirname(sys.executable), 'Main')
    else:
        # Normal development environment
        return os.path.dirname(os.path.abspath(__file__))

def load_config(config_path: str = None) -> Dict:
    """Load configuration from JSON file"""
    # Get possible config file paths to try
    possible_paths = []
    
    # If a specific path was provided, try that first
    if config_path:
        possible_paths.append(config_path)
    
    # Add standard paths to try
    base_path = get_base_path()
    possible_paths.extend([
        os.path.join(base_path, 'config.json'),
        os.path.join(os.path.dirname(base_path), 'config.json'),
        os.path.join(os.getcwd(), 'config.json'),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
    ])
    
    # Try each possible path
    for path in possible_paths:
        print(f"Trying to load config from: {path}")
        if os.path.exists(path):
            try:
                with open(path, 'r') as f:
                    config = json.load(f)
                
                # Update template paths to be absolute
                base_dir = os.path.dirname(os.path.abspath(path))
                for template in config.get('files', {}).values():
                    if 'path' in template:
                        # If path is relative, make it absolute relative to the config file
                        if not os.path.isabs(template['path']):
                            template['path'] = os.path.join(base_dir, template['path'])
                        print(f"Template path: {template['path']}")
                
                print("Successfully loaded config file")
                return config
                
            except json.JSONDecodeError as e:
                print(f"Error: Invalid JSON in {path}: {str(e)}")
            except Exception as e:
                print(f"Error reading {path}: {str(e)}")
    
    print("Error: Could not find or load config file from any standard location")
    return {}

def get_cell_value(workbook: Any, cell_ref: str) -> str:
    """Get value from a cell reference in the workbook"""
    try:
        # Split the reference into sheet name and cell (if sheet is specified)
        if '!' in cell_ref:
            sheet_name, cell = cell_ref.split('!')
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
            cell = cell_ref
        return str(sheet[cell].value)
    except Exception as e:
        print(f"Error reading cell {cell_ref}: {str(e)}")
        return ""

def create_template(excel_path: str, config: Dict[str, str]) -> Dict[str, Any]:
    """
    Create template from Excel file using configuration
    
    Args:
        excel_path: Path to the Excel file
        config: Dictionary mapping cell references to their values
    
    Returns:
        Dictionary containing the template data, mapping coordinates to user responses
    """
    try:
        # Load workbook
        workbook = load_workbook(excel_path, data_only=True)
        
        # Initialize results dictionary
        results: Dict[str, str] = {}
        
        # Process each mapping
        for key, value in config.items():
            # Get the question text from the key cell
            question = get_cell_value(workbook, key)
            # Get the coordinate that will be used as the key
            coordinate = value
            
            # Get user input
            user_input = input(f"{question}: ")
            
            # Store the user input using the coordinate as the key
            results[coordinate] = user_input
            
        return results
        
    except Exception as e:
        print(f"Error creating template: {str(e)}")
        return {}

def display_available_templates(config: Dict) -> None:
    """Display available templates to the user"""
    print("Available templates:")
    for i, template_name in enumerate(config['files'].keys(), 1):
        print(f"{i}. {template_name}")

def get_template_selection(config: Dict) -> str:
    """Get user's template selection"""
    while True:
        display_available_templates(config)
        try:
            choice = int(input("\nSelect a template (enter number): "))
            if 1 <= choice <= len(config['files']):
                return list(config['files'].keys())[choice - 1]
            print("Invalid selection. Please try again.")
        except ValueError:
            print("Please enter a number.")

def save_modified_template(template_path: str, results: Dict[str, str], template_name: str, output_path: str = None) -> str:
    """
    Save a modified copy of the template with the given results.
    
    Args:
        template_path: Path to the template file
        results: Dictionary of cell references and their new values
        template_name: Name of the template (used for generating output filename)
        output_path: Optional custom output path
        
    Returns:
        Path to the saved file
    """
    # Ensure template path is absolute
    if not os.path.isabs(template_path):
        template_path = os.path.join(get_base_path(), template_path)
    
    print(f"Using template: {template_path}")
    
    # Create output directory if it doesn't exist
    if not output_path:
        # Try to create Results directory in the same directory as the executable
        if getattr(sys, 'frozen', False):
            # In PyInstaller bundle
            output_dir = os.path.join(os.path.dirname(sys.executable), 'Results')
        else:
            # In development
            output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Results')
        
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate output filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"{template_name.replace(' ', '_')}_{timestamp}.xlsx"
        output_path = os.path.abspath(os.path.join(output_dir, output_filename))
    
    print(f"Saving output to: {output_path}")
    
    try:
        # Make sure the template exists
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        # Ensure the output directory exists
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        
        # Copy the template to the output path
        shutil.copy2(template_path, output_path)
        
        # Load the workbook and update the specified cells
        wb = load_workbook(output_path)
        
        for coordinate, value in results.items():
            if not value:  # Skip empty values
                continue
                
            try:
                # Handle sheet references (e.g., 'Sheet1!A1')
                if '!' in coordinate:
                    sheet_name, cell_ref = coordinate.split('!')
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                    cell_ref = coordinate
                
                # Update the cell value
                ws[cell_ref] = value
                
            except Exception as e:
                print(f"Warning: Could not update cell {coordinate}: {str(e)}")
        
        # Save the modified workbook
        wb.save(output_path)
        
        print(f"Successfully saved modified template to: {output_path}")
        return output_path
        
    except FileNotFoundError as e:
        print(f"Error: {str(e)}")
        raise
    except PermissionError as e:
        print(f"Permission error: Cannot write to {output_path}. Please check your permissions.")
        raise
    except Exception as e:
        print(f"An unexpected error occurred while saving the template: {str(e)}")
        # Clean up partially written file if it exists
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except Exception as cleanup_error:
                print(f"Warning: Could not clean up temporary file {output_path}: {str(cleanup_error)}")
        raise

def main():
    try:
        print("Starting Template Editor...")
        
        # Load configuration
        config = load_config()  # Will try multiple locations automatically
        
        if not config:
            print("Error: Could not load configuration. Please ensure config.json exists in one of these locations:")
            print(f"  - {os.path.join(get_base_path(), 'config.json')}")
            print(f"  - {os.path.join(os.path.dirname(get_base_path()), 'config.json')}")
            print(f"  - {os.path.join(os.getcwd(), 'config.json')}")
            return 1
            
        if 'files' not in config or not config['files']:
            print("Error: No templates found in configuration.")
            return 1
        
        # Get template selection
        template_name = get_template_selection(config)
        template_config = config['files'].get(template_name)
        
        if not template_config:
            print(f"Error: Configuration for template '{template_name}' not found.")
            return 1
            
        # Get full path to Excel file
        excel_path = template_config.get('path')
        if not excel_path:
            print(f"Error: No path specified for template '{template_name}'")
            return 1
            
        # Ensure the path is absolute
        if not os.path.isabs(excel_path):
            excel_path = os.path.join(get_base_path(), excel_path)
        
        print(f"\nUsing template: {excel_path}")
        
        # Create template
        results = create_template(excel_path, template_config.get('mappings', {}))
        
        if not results:
            print("No results to save. Exiting.")
            return 1
            
        # Save modified template
        try:
            saved_path = save_modified_template(excel_path, results, template_name)
            
            if saved_path and os.path.exists(saved_path):
                print(f"\n✅ Template created successfully!")
                print(f"   Saved to: {saved_path}")
                print("\nUpdated values:")
                for coord, value in results.items():
                    print(f"  {coord}: {value}")
                return 0
            else:
                print("\n❌ Error: Could not save the modified template.")
                return 1
                
        except Exception as e:
            print(f"\n❌ Error saving template: {str(e)}")
            return 1
            
    except KeyboardInterrupt:
        print("\nOperation cancelled by user.")
        return 1
    except Exception as e:
        print(f"\n❌ An unexpected error occurred: {str(e)}")
        return 1

if __name__ == "__main__":
    import sys
    sys.exit(main())
